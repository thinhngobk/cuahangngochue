import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Tạo thư mục nếu chưa có
os.makedirs('static/templates', exist_ok=True)

# Tạo workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mẫu Import Sản Phẩm"

# Header - thêm tonkho
headers = ['tensanpham', 'donvitinh', 'dongiagoc', 'dongiaban', 'tonkho', 'barcode', 'ghichu']
ws.append(headers)

# Style header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Thêm dữ liệu mẫu - thêm tonkho
sample_data = [
    ['Coca Cola 330ml', 'Lon', 8000, 10000, 100, '8934588012345', 'Nước ngọt'],
    ['Pepsi 330ml', 'Lon', 7500, 9500, 50, '8934588054321', ''],
    ['Sting dâu 330ml', 'Lon', 8500, 11000, 75, '', 'Nước tăng lực'],
]

for row in sample_data:
    ws.append(row)

# Điều chỉnh độ rộng cột
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 10  # tonkho
ws.column_dimensions['F'].width = 18  # barcode
ws.column_dimensions['G'].width = 20  # ghichu

# Lưu file
wb.save('static/templates/mau_import_sanpham.xlsx')
print("✅ Đã tạo file template: static/templates/mau_import_sanpham.xlsx")